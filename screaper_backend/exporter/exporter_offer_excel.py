"""
    Creates an excel file which is in the format of an offer
"""

import datetime
import random
from io import BytesIO
from tempfile import NamedTemporaryFile

import numpy as np
from copy import copy

from openpyxl import load_workbook

import os

from dotenv import load_dotenv
from werkzeug.wsgi import FileWrapper

load_dotenv()

class ExporterOfferExcel:

    def _load_template(self):
        filepath = os.getenv('UNSP_TEKLIF_TEMPLATE_PATH')
        print("filepath is: ", filepath)
        workbook = load_workbook(filepath)
        print("Sheetnames are: ", workbook.sheetnames)
        sheet = workbook.active
        print("Workbook is", workbook)
        print("Sheet is", sheet)
        print("Cell F11 is", sheet['F11'])
        print("Cell F11 is", sheet['F11'].internal_value)
        assert sheet['F11'].internal_value == "Ref:", (
            f"Correct cell not found! expected {sheet['F11'].internal_value} got {'Ref'}"
        )

        return workbook, sheet

    def __init__(self):
        self.workbook, self.sheet = self._load_template()
        # set the counter ...
        self.rowcounter = 0
        self.rowoffset = 19  # the enumeration of items starts at row 17

    def update_date(self):
        style = copy(self.sheet[f'H10']._style)
        self.sheet[f'H10'] = '{}'.format(datetime.date.today().strftime("%d.%m.%Y"))
        self.sheet[f'H10']._style = style
        style = copy(self.sheet[f'H11']._style)
        self.sheet[f'H11'] = '{}'.format(datetime.date.today().strftime("%d.%m.%Y"))
        self.sheet[f'H11']._style = style

    def insert_customer(self, customer_obj):
        style = copy(self.sheet[f'A10']._style)
        self.sheet[f'A10'] = '{}'.format(customer_obj.company_name)
        self.sheet[f'A10']._style = style
        style = copy(self.sheet[f'A11']._style)
        self.sheet[f'A11'] = '{}'.format(customer_obj.city)
        self.sheet[f'A11']._style = style
        style = copy(self.sheet[f'A13']._style)
        self.sheet[f'A13'] = 'Tel: {}'.format(customer_obj.phone_number)
        self.sheet[f'A13']._style = style
        style = copy(self.sheet[f'A15']._style)
        self.sheet[f'A15'] = '{}'.format(customer_obj.contact_name if customer_obj.contact_name else "")
        self.sheet[f'A15']._style = style

        style = copy(self.sheet[f'E13']._style)
        self.sheet[f'E13'] = 'Fax: {}'.format(customer_obj.fax_number)
        self.sheet[f'E13']._style = style

        style = copy(self.sheet[f'A15']._style)
        self.sheet[f'F15'] = 'E-mail: {}'.format(customer_obj.email if customer_obj.email else "")
        self.sheet[f'F15']._style = style

    # def insert_reference(self, reference):
    #     style = copy(self.sheet[f'A10']._style)
    #     self.sheet[f'A10'] = '{}'.format(reference)
    #     self.sheet[f'A10']._style = style

    def insert_item(
            self,
            description,
            partnumber,
            listprice,
            requested_units,
            margin_multiplier=2.5,
            stock=None,
            status=None,
            weight=None,
            replaced=None
    ):

        assert isinstance(listprice, float) or isinstance(listprice, int) or isinstance(listprice, np.float64), (
            "Listprice is not of type float", listprice, type(listprice)
        )

        rowidx = self.rowcounter + self.rowoffset

        # Insert a row
        if self.rowcounter > 0:
            self.sheet.insert_rows(rowidx + 1) # Plus 1 because we insert above, not below the roindex
            previous_row = rowidx - 1
        else:
            previous_row = rowidx

        # Insert "sira"
        style = copy(self.sheet[f'A{previous_row}']._style)
        self.sheet[f'A{rowidx}'] = f'=A{previous_row} + 1' if self.rowcounter > 0 else 1
        self.sheet[f'A{rowidx}']._style = style

        # Reference no.
        style = copy(self.sheet[f'C{previous_row}']._style)
        self.sheet[f'C{rowidx}'] = f'=C{previous_row} + 1' if self.rowcounter > 0 else 1
        self.sheet[f'C{rowidx}']._style = style

        # Partnumber
        style = copy(self.sheet[f'D{previous_row}']._style)
        self.sheet[f'D{rowidx}'] = partnumber
        self.sheet[f'D{rowidx}']._style = style

        # Description
        style = copy(self.sheet[f'E{previous_row}']._style)
        self.sheet[f'E{rowidx}'] = description
        self.sheet[f'E{rowidx}']._style = style

        # Buy price
        style = copy(self.sheet[f'I{previous_row}']._style)
        self.sheet[f'I{rowidx}'] = listprice
        self.sheet[f'I{rowidx}']._style = style

        # Number Units
        style = copy(self.sheet[f'B{previous_row}']._style)
        self.sheet[f'B{rowidx}'] = requested_units
        self.sheet[f'B{rowidx}']._style = style

        # Stock
        if stock is not None:
            style = copy(self.sheet[f'L{previous_row}']._style)
            self.sheet[f'L{rowidx}'] = stock
            self.sheet[f'L{rowidx}']._style = style

        # Status
        if status is not None:
            style = copy(self.sheet[f'K{previous_row}']._style)
            self.sheet[f'K{rowidx}'] = status
            self.sheet[f'K{rowidx}']._style = style

        # Weight
        if weight is not None:
            style = copy(self.sheet[f'M{previous_row}']._style)
            self.sheet[f'M{rowidx}'] = weight
            self.sheet[f'M{rowidx}']._style = style

        # Replaced
        if replaced is not None:
            style = copy(self.sheet[f'N{previous_row}']._style)
            self.sheet[f'N{rowidx}'] = replaced
            self.sheet[f'N{rowidx}']._style = style

        # Copy all equations which were not copied yet
        style = copy(self.sheet[f'F{previous_row}']._style)
        self.sheet[f'F{rowidx}'] = f'=J{rowidx}'
        self.sheet[f'F{rowidx}']._style = style

        style = copy(self.sheet[f'J{previous_row}']._style)
        self.sheet[f'J{rowidx}'] = f'=I{rowidx} * {margin_multiplier}'
        self.sheet[f'J{rowidx}']._style = style

        style = copy(self.sheet[f'H{previous_row}']._style)
        self.sheet[f'H{rowidx}'] = f'=F{rowidx}*B{rowidx}'
        self.sheet[f'H{rowidx}']._style = style

        # Copy cell style for "dead cells
        for deadcol in ['C', 'G']:
            style = copy(self.sheet[f'{deadcol}{previous_row}']._style)
            self.sheet[f'{deadcol}{rowidx}']._style = style

        self.sheet[f'H{rowidx + 3}'] = f'=SUM(H{self.rowoffset}: H{rowidx})'
        self.sheet[f'H{rowidx + 4}'] = f'=H{rowidx + 3}*15/100'
        self.sheet[f'H{rowidx + 5}'] = f'=H{rowidx + 3}-H{rowidx + 4}'

        # Increase counter by one...
        self.rowcounter += 1

    def save_to_disk(self):
        rnd_no = random.randint(1000, 9999)
        self.workbook.save(f"./test{rnd_no}.xlsx")

    def save_to_disk_from_bytes(self):
        """
            Checks if byteencoding makes it unreadable or not
        :return:
        """
        rnd_no = random.randint(1000, 9999)
        with NamedTemporaryFile() as tmp:
            with open(f"./test{rnd_no}.xlsx", "wb") as fp:
                self.workbook.save(tmp.name)

                output = BytesIO(tmp.read())
                # Write to tmp file
                fp.write(output.read())

    def get_bytestring(self):
        # Have a look at this for better excel exporting: https://flask-excel.readthedocs.io/en/v0.0.7/

        with NamedTemporaryFile() as tmp:
            self.workbook.save(tmp.name)
            output = BytesIO(tmp.read())
            wrapped_file = FileWrapper(output)
            return wrapped_file

        # print("tmp is: ", tmp)
        # return tmp
        # output = BytesIO(tmp.read())
        # return output.read()
